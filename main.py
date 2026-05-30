from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

LABELS = ("SELF_CONTAINED", "NEEDS_PRIOR_DEF", "INFORMAL")


@dataclass(frozen=True)
class SourceText:
    path: Path
    source_kind: str
    content: str


@dataclass(frozen=True)
class Candidate:
    item_id: str
    definition_ru: str
    source_file: str
    source_kind: str


@dataclass(frozen=True)
class DatasetRecord:
    id: str
    term: str
    definition_ru: str
    source_file: str
    source_kind: str
    predicted_label: str
    formalizable_now: str
    rule_score: float
    evidence: str
    label_reason: str
    missing_context: str


class TextNormalizer:
    @staticmethod
    def clean(text: str) -> str:
        text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
        text = text.replace("\u00ad", "")
        text = re.sub(r"-\s*\n\s*", "", text)
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text.strip()


class SourceCollector:
    def __init__(self, source_dir: Path):
        self.text_dir = source_dir / "texts"
        self.pdf_dir = source_dir / "pdf"

    def collect(self) -> list[SourceText]:
        documents: list[SourceText] = []

        if self.text_dir.exists():
            for path in sorted(self.text_dir.glob("*.txt")):
                content = TextNormalizer.clean(path.read_text(encoding="utf-8"))
                documents.append(SourceText(path, "txt", content))

        if self.pdf_dir.exists():
            for path in sorted(self.pdf_dir.glob("*.pdf")):
                content = TextNormalizer.clean(self._read_pdf(path))
                documents.append(SourceText(path, "pdf", content))

        return documents

    @staticmethod
    def _read_pdf(path: Path) -> str:
        try:
            import pymupdf
        except ModuleNotFoundError as exc:
            raise RuntimeError("Для чтения PDF установите зависимость: pip install PyMuPDF") from exc

        parts: list[str] = []

        with pymupdf.open(path) as document:
            for page in document:
                parts.append(page.get_text("text", sort=True))

        return "\n".join(parts)


class DefinitionExtractor:
    block_pattern = re.compile(r"\[(D\d{3})\]\s*(.*?)(?=(?:\n\s*\[D\d{3}\])|\Z)", re.DOTALL)
    sentence_pattern = re.compile(r"(?<=[.!?])\s+")
    definition_marker = re.compile(
        r"(?i)(определени[ея]|называ(?:ется|ются)|будем называть|назов[её]м|defined as|is called)"
    )

    def extract(self, documents: Iterable[SourceText]) -> list[Candidate]:
        records: list[Candidate] = []
        generated_id = 1

        for document in documents:
            blocks = list(self.block_pattern.finditer(document.content))

            if blocks:
                for match in blocks:
                    text = TextNormalizer.clean(match.group(2))
                    records.append(Candidate(match.group(1), text, document.path.name, document.source_kind))
                continue

            for sentence in self.sentence_pattern.split(document.content):
                sentence = TextNormalizer.clean(sentence)

                if len(sentence) >= 35 and self.definition_marker.search(sentence):
                    records.append(Candidate(f"P{generated_id:03d}", sentence, document.path.name, document.source_kind))
                    generated_id += 1

        return sorted(records, key=lambda value: value.item_id)


class TermExtractor:
    patterns = (
        re.compile(r"^(.{2,80}?)\s+называ(?:ется|ются)\s+", re.IGNORECASE),
        re.compile(r"^(.{2,80}?)\s+будем называть\s+", re.IGNORECASE),
        re.compile(r"^(.{2,80}?)\s+назов[её]м\s+", re.IGNORECASE),
        re.compile(r"^определени[ея]\s+(.{2,80}?)[—-]", re.IGNORECASE),
    )

    @classmethod
    def extract(cls, text: str) -> str:
        normalized = re.sub(r"\s+", " ", text).strip()

        for pattern in cls.patterns:
            match = pattern.search(normalized)

            if match:
                term = match.group(1).strip(" .,:;—-")
                return term[:80]

        return ""


class RuleAnnotator:
    dependent_terms = {
        "предыдущ": "ссылка на предыдущий фрагмент",
        "ранее": "ссылка на ранее заданный объект или условие",
        "введённ": "используется ранее введённый объект",
        "введенн": "используется ранее введённый объект",
        "выше": "ссылка на внешнее описание",
        "сохраняя обозначения": "не заданы используемые обозначения",
        "данной категории": "категория не описана в формулировке",
        "согласно выбранному": "используется ранее выбранная конструкция",
        "соглашениях раздела": "опора на соглашения вне определения",
        "зафиксированной ранее": "используется зафиксированный ранее объект",
        "обозначенного выше": "объект определён вне фрагмента",
    }

    vague_terms = {
        "неформальн": "явная отметка неформальности",
        "немного": "отсутствует числовая граница",
        "несуществен": "не задана допустимая величина изменения",
        "достаточно быстро": "не задан критерий скорости",
        "визуальн": "критерий зависит от восприятия",
        "много": "не задан порог количества",
        "легко": "не определена мера сложности",
        "незначитель": "не задана погрешность",
        "заметн": "оценочный признак",
        "почти всегда": "не задана вероятность или доля",
        "тесно": "не задано расстояние",
        "не слишком": "не задано ограничение",
        "очень редко": "не задана вероятность",
        "мало шагов": "не задан порог числа шагов",
        "особых свойств": "не перечислены свойства",
        "напоминает": "нет формального отношения",
        "пренебречь": "не задан допуск",
        "выглядит логично": "субъективная оценка",
        "серьёзно влияет": "не задана мера влияния",
    }

    formal_cues = (
        r"\bесли\b",
        r"\bдля любого\b",
        r"\bдля каждого\b",
        r"\bсуществует\b",
        r"тогда и только тогда",
        r"∈",
        r"=",
        r"≤",
        r"≥",
        r"<",
        r">",
    )

    def annotate(self, text: str) -> tuple[str, str, str, str, float, str]:
        lower = text.lower()

        vague_hits = [f"{term}: {reason}" for term, reason in self.vague_terms.items() if term in lower]

        if vague_hits:
            evidence = "; ".join(vague_hits)
            return (
                "INFORMAL",
                "Нет",
                "Обнаружены качественные или субъективные критерии без точного условия.",
                "Следует заменить оценочные слова числовым либо логическим условием.",
                0.96,
                evidence,
            )

        dependent_hits = [f"{term}: {reason}" for term, reason in self.dependent_terms.items() if term in lower]

        if dependent_hits:
            evidence = "; ".join(dependent_hits)
            return (
                "NEEDS_PRIOR_DEF",
                "После восстановления контекста",
                "Формулировка зависит от обозначений или объектов, заданных вне найденного фрагмента.",
                "Необходимо добавить определения используемых объектов и условий.",
                0.93,
                evidence,
            )

        cues = [pattern for pattern in self.formal_cues if re.search(pattern, lower)]

        if cues:
            evidence = "формальные маркеры: " + ", ".join(cues)
            return (
                "SELF_CONTAINED",
                "Да",
                "Объект и проверяемое условие заданы непосредственно в формулировке.",
                "Не требуется.",
                0.94,
                evidence,
            )

        return (
            "NEEDS_PRIOR_DEF",
            "После уточнения",
            "Не найдено явное проверяемое условие, достаточное для формальной записи.",
            "Требуется уточнить объект и критерий определения.",
            0.60,
            "формальные маркеры не обнаружены",
        )


class DatasetAssembler:
    def __init__(self):
        self.annotator = RuleAnnotator()

    def assemble(self, candidates: list[Candidate]) -> list[DatasetRecord]:
        records: list[DatasetRecord] = []

        for candidate in candidates:
            label, formalizable, reason, missing, score, evidence = self.annotator.annotate(candidate.definition_ru)

            records.append(
                DatasetRecord(
                    id=candidate.item_id,
                    term=TermExtractor.extract(candidate.definition_ru),
                    definition_ru=candidate.definition_ru,
                    source_file=candidate.source_file,
                    source_kind=candidate.source_kind,
                    predicted_label=label,
                    formalizable_now=formalizable,
                    rule_score=score,
                    evidence=evidence,
                    label_reason=reason,
                    missing_context=missing,
                )
            )

        return records


class Exporter:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write_json(self, records: list[DatasetRecord]) -> Path:
        path = self.output_dir / "definitions_dataset.json"
        path.write_text(
            json.dumps([asdict(record) for record in records], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return path

    def write_xlsx(self, records: list[DatasetRecord]) -> Path:
        if not records:
            raise RuntimeError("Нет данных для записи в Excel.")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ModuleNotFoundError as exc:
            raise RuntimeError("Для формирования XLSX установите зависимость: pip install openpyxl") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Definitions"

        headers = list(asdict(records[0]).keys())
        sheet.append(headers)

        for record in records:
            data = asdict(record)
            sheet.append([data[header] for header in headers])

        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        widths = {
            "A": 12,
            "B": 28,
            "C": 80,
            "D": 30,
            "E": 16,
            "F": 24,
            "G": 22,
            "H": 14,
            "I": 48,
            "J": 55,
            "K": 55,
        }

        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        sheet.freeze_panes = "A2"

        end_row = len(records) + 1
        end_column = get_column_letter(len(headers))
        table = Table(displayName="DefinitionsTable", ref=f"A1:{end_column}{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        sheet.add_table(table)

        path = self.output_dir / "Definitions_Review.xlsx"
        workbook.save(path)

        return path

    def write_summary(self, records: list[DatasetRecord], documents: list[SourceText], excel_path: Path, json_path: Path) -> Path:
        distribution = Counter(record.predicted_label for record in records)

        summary = {
            "theme": "Определения: полнота и формальная записываемость",
            "records_total": len(records),
            "source_files": {
                "txt": sum(document.source_kind == "txt" for document in documents),
                "pdf": sum(document.source_kind == "pdf" for document in documents),
            },
            "class_distribution": dict(distribution),
            "output": {
                "excel": str(excel_path),
                "json": str(json_path),
            },
        }

        path = self.output_dir / "run_summary.json"
        path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

        return path


class Application:
    def __init__(self, root: Path):
        self.root = root

    def build(self) -> dict[str, object]:
        documents = SourceCollector(self.root / "sources").collect()

        if not documents:
            raise RuntimeError("Исходные файлы не найдены. Добавьте .txt в sources/texts или .pdf в sources/pdf.")

        candidates = DefinitionExtractor().extract(documents)

        if not candidates:
            raise RuntimeError("Определения не найдены. Используйте формат [D001] или текст со словами 'называется', 'определение', 'defined as'.")

        records = DatasetAssembler().assemble(candidates)

        exporter = Exporter(self.root / "output")
        json_path = exporter.write_json(records)
        excel_path = exporter.write_xlsx(records)
        summary_path = exporter.write_summary(records, documents, excel_path, json_path)

        return {
            "status": "ok",
            "records_total": len(records),
            "excel_file": str(excel_path),
            "json_file": str(json_path),
            "summary_file": str(summary_path),
        }


def main() -> None:
    parser = argparse.ArgumentParser(description="Извлечение определений из TXT/PDF и выгрузка результата в Excel.")
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parent)

    args = parser.parse_args()
    result = Application(args.root).build()

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()